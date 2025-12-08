import openpyxl

print("--- Invoice Generator Start ---")

# 1. 新しいExcelブックを作る
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Seikyu-sho"  # シートの名前

# 2. 請求データ（商品名, 個数, 単価）
items = [
    ["Apple", 3, 100],
    ["Pen", 5, 200],
    ["Book", 2, 1500]
]

# 3. 見出し（ヘッダー）を書く
sheet["A1"] = "請求書 (Invoice)"
sheet["A3"] = "商品名"
sheet["B3"] = "個数"
sheet["C3"] = "単価"
sheet["D3"] = "金額"

# 4. データをループして書き込む
current_row = 4  # 4行目からスタート
total_amount = 0

for item in items:
    name = item[0]
    count = item[1]
    price = item[2]
    subtotal = count * price  # 小計を計算
    
    # セルに書き込む (row=行, column=列)
    sheet.cell(row=current_row, column=1, value=name)
    sheet.cell(row=current_row, column=2, value=count)
    sheet.cell(row=current_row, column=3, value=price)
    sheet.cell(row=current_row, column=4, value=subtotal)
    
    total_amount = total_amount + subtotal
    current_row = current_row + 1

# 5. 最後に合計を書く
# f文字列を使って、今の行（current_row）を指定するテクニック
sheet[f"C{current_row + 1}"] = "合計金額"
sheet[f"D{current_row + 1}"] = total_amount

# 6. ファイルとして保存！
wb.save("invoice.xlsx")
print("Success! Created 'invoice.xlsx'")
