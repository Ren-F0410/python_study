import openpyxl

print("--- Excel Reader Start ---")

# 1. さっき作ったファイルを開く（読み込みモード）
# data_only=True にすると、計算式ではなく「計算結果の数字」を読み込めます
wb = openpyxl.load_workbook("invoice.xlsx", data_only=True)

# 2. シートを選ぶ
sheet = wb["Seikyu-sho"]

# 3. 指定したセルの中身を読む
# C7の「合計金額」という文字と、D7の「4300」という数字を読みます
label = sheet["C8"].value
amount = sheet["D8"].value

print("読み取ったデータ:")
print(f"{label} は {amount} 円です。")

# 4. 行を順番に全部読んでみる（分析の基礎！）
print("-" * 20)
print("詳細リスト:")

# 4行目から、データがある最後の行（max_row）までループ
for i in range(4, sheet.max_row):
    # A列(1)の商品名と、D列(4)の金額を取得
    name = sheet.cell(row=i, column=1).value
    price = sheet.cell(row=i, column=4).value
    print(f"- {name}: {price}円")

print("--- Analysis Complete! ---")
